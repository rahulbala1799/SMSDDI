import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np

# Region Mapping Dictionary
region_mapping = {
    "United States": "US",
    "Finland": "Finland",
    "United Kingdom(Mainland)": "UK",
    "Ireland(Rep.)": "Ireland",
    "Germany": "Germany",
    "Jersey": "UK",
    "United Kingdom(Northern Ireland)": "UK",
    "United Kingdom(NI)": "UK",
    "Canada": "US",
    "United Arab Emirates": "Ireland",
    "Australia": "Australia",
    "Bermuda": "Ireland",
    "Guernsey": "UK",
    "Switzerland": "Germany",
    "Austria": "Germany",
    "India": "Ireland",
    "Bahrain": "Ireland",
    "Puerto Rico": "US",
    "New Caledonia": "Australia",
    "South Africa": "Ireland",
    "Spain": "Germany",
    "Guatemala": "Ireland",
    "Luxembourg": "Ireland",
    "Netherlands Antilles": "US",
    "New Zealand": "Australia",
    "Gibraltar": "Ireland",
    "Mauritius": "Ireland",
    "Netherlands": "Ireland",
    "Sweden": "Ireland",
    "Malta": "Ireland",
    "France": "Ireland",
    "Isle of Man": "UK",
    "Martinique": "Ireland",
    "Seychelles": "Ireland",
    "Cayman Islands": "US",
    "Saudi Arabia": "Ireland",
    "Pakistan": "Ireland",
}

# Page Configuration
st.set_page_config(
    page_title="Processing Page",
    page_icon="⚙️",
    layout="wide",
)

st.title("Processing Page")

# Check if the file is available from step 1
if "uploaded_file_path" in st.session_state and st.session_state["uploaded_file_path"]:
    st.success("File successfully loaded for processing!")

    # Load the uploaded file
    file_path = st.session_state["uploaded_file_path"]
    try:
        # Load the data into a Pandas DataFrame
        df = pd.read_excel(file_path)

        # Filter rows where LINE ITEM contains "SMS" or "SMS Bundle Sales"
        sms_df = df[df["LINE ITEM"].str.strip().str.upper().isin(["SMS", "SMS BUNDLE SALES"])]

        # Update SMS TYPE column for SMS Bundle Sales
        sms_df["SMS TYPE"] = np.where(
            sms_df["LINE ITEM"].str.strip().str.upper() == "SMS BUNDLE SALES",
            "Bundle/Purchase",
            sms_df["SMS TYPE"]
        )

        # Create the Type column based on SMS TYPE
        sms_df["Type"] = np.where(
            sms_df["SMS TYPE"].str.lower() == "usage",
            "Usage",
            "Bundle/Purchase"
        )

        # Add a column for Region based on LOCATION
        sms_df["Region"] = sms_df["LOCATION"].map(region_mapping).fillna("Unknown")

        # Calculate Totals
        total_amount = sms_df["AMOUNT"].sum()
        usage_total = sms_df.loc[sms_df["Type"] == "Usage", "AMOUNT"].sum()
        bundle_total = sms_df.loc[sms_df["Type"] == "Bundle/Purchase", "AMOUNT"].sum()

        # Display totals dynamically
        st.write("### Totals:")
        st.write(f"- **Total Amount:** {total_amount}")
        st.write(f"- **Usage Total:** {usage_total}")
        st.write(f"- **Bundle/Purchase Total:** {bundle_total}")

        # Load the existing workbook and create a new tab
        wb = load_workbook(file_path)
        sms_tab_name = "Processed_SMS"

        if sms_tab_name in wb.sheetnames:
            del wb[sms_tab_name]  # Ensure the tab is replaced if it already exists
        ws = wb.create_sheet(sms_tab_name)

        # Write the DataFrame to the new tab with styled headers and alternating row colors
        header_font = Font(bold=True)
        light_gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        for r_idx, row in enumerate(dataframe_to_rows(sms_df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Header row
                    cell.font = header_font
                else:  # Alternate row colors
                    cell.fill = light_gray_fill if r_idx % 2 == 0 else white_fill

        # Save the updated workbook
        updated_file_path = "processed_file.xlsx"
        wb.save(updated_file_path)

        st.success("Data processed successfully and added to a new tab in the Excel file!")
        st.write("### Processed SMS Data:")
        st.dataframe(sms_df)

        # Download the updated Excel file
        with open(updated_file_path, "rb") as file:
            st.download_button(
                label="Download Updated Excel File",
                data=file,
                file_name="processed_file.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    except Exception as e:
        st.error(f"An error occurred while processing the file: {e}")
else:
    st.warning("No file uploaded. Please upload a file on the main page.")
