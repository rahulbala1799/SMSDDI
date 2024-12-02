import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
from datetime import date, timedelta
import difflib

# Page Configuration
st.set_page_config(
    page_title="DDI and Plan Journals Creation",
    page_icon="📄",
    layout="wide",
)

st.title("DDI and Plan Journals Creation")

# File Upload
uploaded_file = st.file_uploader("Upload the processed Excel file from Step 2:", type=["xlsx"])

# Calendar widget for journal date
journal_date = st.date_input("Select the journal date:")
selected_month = journal_date.strftime("%B")
reversal_date = (journal_date.replace(day=1) + timedelta(days=32)).replace(day=2)

# Mapping table
location_mapping = {
    "Australia": "Australia",
    "Austria": "Germany : Austria",
    "Bahrain": "Ireland : ROW : Bahrain",
    "Bermuda": "North America : Bermuda",
    "Canada": "North America : Canada",
    "Cayman Islands": "North America : Cayman Islands",
    "Cameroon": "Ireland : ROW : Cameroon",
    "Finland": "Finland",
    "France": "Ireland : ROE : France",
    "Germany": "Germany",
    "Guatemala": "Ireland : ROW : Guatemala",
    "Guernsey": "UK : Guernsey",
    "India": "Ireland : ROW : India",
    "Ireland(Rep.)": "Ireland : Ireland(Rep.)",
    "Isle of Man": "UK : Isle of Man",
    "Jersey": "UK : Jersey",
    "Luxembourg": "Ireland : ROE : Luxembourg",
    "Martinique": "Ireland : ROE : France : Martinique",
    "Malta": "Ireland : ROE : Malta",
    "Netherlands": "Netherlands",
    "Netherlands Antilles": "North America : Netherlands Antilles",
    "New Caledonia": "Australia : New Caledonia",
    "Puerto Rico": "North America : Puerto Rico",
    "Seychelles": "Ireland : ROW : Seychelles",
    "South Africa": "Ireland : ROW : South Africa",
    "Spain": "Germany : Spain",
    "Sweden": "Ireland : ROE : Sweden",
    "Switzerland": "Germany : Switzerland",
    "United Arab Emirates": "United Arab Emirates",
    "United Kingdom(Mainland)": "UK : United Kingdom(Mainland)",
    "United Kingdom(NI)": "UK : United Kingdom(NI)",
    "United States": "North America : United States",
    "Saudi Arabia": "Ireland : ROW : Saudi Arabia",
    "New Zealand": "Australia : New Zealand",
    "Kuwait": "Ireland : ROW : United Arab Emirates",
    "Gibraltar": "Ireland : ROW : Gibraltar",
    "Mauritius": "Ireland : ROW : Mauritius",
    "Qatar": "Ireland : ROW : Qatar",
    "Trinidad and Tobago": "Ireland : ROW : Trinidad and Tobago",
    "Anguilla": "Ireland : ROW : Anguilla",
    "Cyprus": "Ireland : ROE : Cyprus",
    "United Kingdom(Northern Ireland)": "UK : United Kingdom(NI)",
    "Pakistan": "Ireland : ROW : Pakistan",
    "Mexico": "Ireland : ROW : Mexico"
}

# Helper function for LOCATION mapping
def map_location(location):
    if location in location_mapping:
        return location_mapping[location]
    else:
        # Use closest match logic
        closest_match = difflib.get_close_matches(location, location_mapping.keys(), n=1)
        return location_mapping[closest_match[0]] if closest_match else "Unknown"

if uploaded_file and journal_date:
    try:
        # Load the workbook
        wb = load_workbook(uploaded_file)
        if "Processed_SMS" not in wb.sheetnames:
            st.error("The 'Processed_SMS' tab is missing. Please upload the correct file.")
        else:
            # Load data from the Processed_SMS tab
            sms_df = pd.read_excel(uploaded_file, sheet_name="Processed_SMS")

            # Ensure required columns are present
            required_columns = {"Subsidiary", "LOCATION", "AMOUNT", "CURRENCY", "CLIENT", "Type"}
            if not required_columns.issubset(sms_df.columns):
                st.error(f"The uploaded file is missing one or more required columns: {', '.join(required_columns)}")
            else:
                # Progress bar and logs
                progress = st.progress(0)
                log_area = st.empty()
                total_rows = len(sms_df)
                log_frequency = max(1, total_rows // 10)  # Update log every ~10% of rows

                # Create DDI Journal tab
                if "DDI Journal" in wb.sheetnames:
                    del wb["DDI Journal"]
                ws_ddi = wb.create_sheet(title="DDI Journal")

                # Create Plan Journals tab
                if "Plan Journals" in wb.sheetnames:
                    del wb["Plan Journals"]
                ws_plan = wb.create_sheet(title="Plan Journals")

                # Write headers for both tabs
                headers = [
                    "Entry No.", "Date", "Subsidiary", "Location", "Currency", "Account", "Memo",
                    "Debit", "Credit", "Department", "Cost Centre", "Mgmt P&L", "Name"
                ]
                for ws in [ws_ddi, ws_plan]:
                    for col_idx, header in enumerate(headers, start=1):
                        cell = ws.cell(row=1, column=col_idx, value=header)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = Border(
                            left=Side(style="thin"), right=Side(style="thin"),
                            top=Side(style="thin"), bottom=Side(style="thin")
                        )

                # Populate DDI Journal entries
                for entry_no, row in enumerate(sms_df.itertuples(index=False), start=1):
                    subsidiary = row.Subsidiary
                    location = map_location(row.LOCATION)
                    currency = row.CURRENCY
                    amount = row.AMOUNT
                    name = row.CLIENT
                    memo = f"{selected_month} SMS DDI"

                    # Odd row (Debit journal)
                    ws_ddi.append([
                        entry_no, journal_date, subsidiary, location, currency,
                        "40040 Sales : SMS", memo, "", amount,
                        "Revenue", "Revenue", "Recurring Revenue", name
                    ])

                    # Even row (Credit journal)
                    ws_ddi.append([
                        entry_no, journal_date, subsidiary, location, currency,
                        "16050 Other Current Assets : Accrued SMS", memo, amount, "",
                        "Balance Sheet", "Balance Sheet", "Balance Sheet", name
                    ])

                # Populate Plan Journals entries (Filtered by Type = Usage)
                filtered_df = sms_df[sms_df["Type"] == "Usage"]
                for entry_no, row in enumerate(filtered_df.itertuples(index=False), start=1):
                    subsidiary = row.Subsidiary
                    location = map_location(row.LOCATION)
                    currency = row.CURRENCY
                    amount = row.AMOUNT
                    name = row.CLIENT
                    memo = f"{selected_month} SMS Plan Usage"

                    # Debit row
                    ws_plan.append([
                        entry_no, journal_date, subsidiary, location, currency,
                        "40040 Sales : SMS", memo, amount, "",
                        "Revenue", "Revenue", "Recurring Revenue", name
                    ])

                    # Credit row
                    ws_plan.append([
                        entry_no, journal_date, subsidiary, location, currency,
                        "40041 Sales : SMS : SMS Plan Usage in Month", memo, "", amount,
                        "Revenue", "Revenue", "Recurring Revenue", name
                    ])

                    # Update logs and progress
                    if entry_no % log_frequency == 0 or entry_no == total_rows:
                        log_area.write(f"Processed {entry_no}/{total_rows} rows for Plan Journals...")
                        progress.progress(entry_no / total_rows)

                # Save the updated workbook
                output = BytesIO()
                wb.save(output)
                output.seek(0)

                # Provide download link
                st.success("DDI and Plan Journals created successfully!")
                st.download_button(
                    label="Download Updated File",
                    data=output,
                    file_name="ddi_and_plan_journals.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

    except Exception as e:
        st.error(f"An error occurred while processing the file: {e}")
else:
    st.info("Please upload the processed file and select a journal date to proceed.")
