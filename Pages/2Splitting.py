import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO

# Page Configuration
st.set_page_config(
    page_title="Summary Table Generator",
    page_icon="📊",
    layout="wide",
)

st.title("Summary Table Generator")

# File Upload
uploaded_file = st.file_uploader("Upload the processed Excel file from Step 2:", type=["xlsx"])

if uploaded_file:
    try:
        # Load the workbook
        wb = load_workbook(uploaded_file)
        if "Processed_SMS" not in wb.sheetnames:
            st.error("The 'Processed_SMS' tab is missing. Please upload the correct file.")
        else:
            # Load data from the Processed_SMS tab
            sms_df = pd.read_excel(uploaded_file, sheet_name="Processed_SMS")

            # Ensure required columns are present
            required_columns = {"Subsidiary", "LOCATION", "AMOUNT", "Type", "CURRENCY"}
            if not required_columns.issubset(sms_df.columns):
                st.error(f"The uploaded file is missing one or more required columns: {', '.join(required_columns)}")
            else:
                # Function to style the worksheet
                def style_worksheet(ws, headers):
                    for col_idx, header in enumerate(headers, start=1):
                        cell = ws.cell(row=1, column=col_idx, value=header)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = Border(
                            left=Side(style="thin"), right=Side(style="thin"),
                            top=Side(style="thin"), bottom=Side(style="thin")
                        )

                # 1. Sales by Subsidiary Tab
                st.header("Processing Sales by Subsidiary...")
                sales_data = sms_df.groupby(["Subsidiary", "Type"])["AMOUNT"].sum().unstack(fill_value=0).reset_index()
                sales_data["Grand Total"] = sales_data.get("Usage", 0) + sales_data.get("Bundle/Purchase", 0)

                # Display breakdown by subsidiary during processing
                for _, row in sales_data.iterrows():
                    st.write(f"Subsidiary: {row['Subsidiary']}, Usage: {row.get('Usage', 0)}, "
                             f"Bundle/Purchase: {row.get('Bundle/Purchase', 0)}, Grand Total: {row['Grand Total']}")

                if "Sales by Subsidiary" in wb.sheetnames:
                    del wb["Sales by Subsidiary"]
                ws_sales = wb.create_sheet(title="Sales by Subsidiary")
                headers = ["Subsidiary", "Usage", "Bundle/Purchase", "Grand Total"]
                style_worksheet(ws_sales, headers)

                for row_idx, row in enumerate(sales_data.itertuples(index=False), start=2):
                    for col_idx, value in enumerate(row, start=1):
                        cell = ws_sales.cell(row=row_idx, column=col_idx, value=value)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = Border(
                            left=Side(style="thin"), right=Side(style="thin"),
                            top=Side(style="thin"), bottom=Side(style="thin")
                        )
                        cell.fill = PatternFill(
                            start_color="D9E1F2" if row_idx % 2 == 0 else "FFFFFF",
                            end_color="D9E1F2" if row_idx % 2 == 0 else "FFFFFF",
                            fill_type="solid"
                        )

                # 2. Overall Plan Usage Tab
                st.header("Processing Overall Plan Usage...")
                usage_data = sms_df[sms_df["Type"] == "Usage"].groupby(
                    ["Subsidiary", "LOCATION", "CURRENCY"]
                )["AMOUNT"].sum().reset_index()

                # Display breakdown by subsidiary and location during processing
                for _, row in usage_data.iterrows():
                    st.write(f"Subsidiary: {row['Subsidiary']}, Location: {row['LOCATION']}, "
                             f"Currency: {row['CURRENCY']}, Amount: {row['AMOUNT']}")

                if "Overall Plan Usage" in wb.sheetnames:
                    del wb["Overall Plan Usage"]
                ws_overall = wb.create_sheet(title="Overall Plan Usage")
                headers = ["Subsidiary", "LOCATION", "CURRENCY", "Sum of AMOUNT"]
                style_worksheet(ws_overall, headers)

                for row_idx, row in enumerate(usage_data.itertuples(index=False), start=2):
                    for col_idx, value in enumerate(row, start=1):
                        cell = ws_overall.cell(row=row_idx, column=col_idx, value=value)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = Border(
                            left=Side(style="thin"), right=Side(style="thin"),
                            top=Side(style="thin"), bottom=Side(style="thin")
                        )
                        cell.fill = PatternFill(
                            start_color="D9E1F2" if row_idx % 2 == 0 else "FFFFFF",
                            end_color="D9E1F2" if row_idx % 2 == 0 else "FFFFFF",
                            fill_type="solid"
                        )

                # 3. Totals Tab
                st.header("Generating Totals Page...")
                totals_data = {
                    "Category": ["Total Usage", "Total Bundle/Purchase", "Grand Total"],
                    "Amount": [
                        sales_data["Usage"].sum(),
                        sales_data["Bundle/Purchase"].sum(),
                        sales_data["Grand Total"].sum(),
                    ],
                }
                totals_df = pd.DataFrame(totals_data)

                if "Totals" in wb.sheetnames:
                    del wb["Totals"]
                ws_totals = wb.create_sheet(title="Totals")
                headers = ["Category", "Amount"]
                style_worksheet(ws_totals, headers)

                for row_idx, row in enumerate(totals_df.itertuples(index=False), start=2):
                    for col_idx, value in enumerate(row, start=1):
                        cell = ws_totals.cell(row=row_idx, column=col_idx, value=value)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = Border(
                            left=Side(style="thin"), right=Side(style="thin"),
                            top=Side(style="thin"), bottom=Side(style="thin")
                        )
                        cell.fill = PatternFill(
                            start_color="D9E1F2" if row_idx % 2 == 0 else "FFFFFF",
                            end_color="D9E1F2" if row_idx % 2 == 0 else "FFFFFF",
                            fill_type="solid"
                        )

                # Save the updated workbook
                output = BytesIO()
                wb.save(output)
                output.seek(0)

                # Provide download link
                st.success("Summary tables added successfully!")
                st.download_button(
                    label="Download Updated Excel File",
                    data=output,
                    file_name="summary_tables_output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

    except Exception as e:
        st.error(f"An error occurred while processing the file: {e}")
else:
    st.info("Please upload the processed file from Step 2 to proceed.")
